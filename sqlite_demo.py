# coding = utf-8

import sqlite3
from openpyxl import load_workbook


def get_excel_data():
    wb = load_workbook(r'D:\Code\Python-Study-Notes\exercise\猫眼热映榜.xlsx')

    films = []

    for row in range(1, (wb['Sheet'].max_row + 1)):
        film = []
        for col in range(1, (wb['Sheet'].max_column + 1)):
            film.append(wb['Sheet'].cell(row=row, column=col).value)
        films.append(film)

    return films


def create_database():
    db_link = sqlite3.connect('sqlite_demo.db')
    db_csor = db_link.cursor()

    db_csor.execute(r'''
            CREATE TABLE FILMS_INFO
            (
                films_names varchar(100),
                protagonist varchar(999),
                releasetime date,
                posterslink text
            )
        ''')

    db_csor.close()
    db_link.close()


def insert_database():
    films = get_excel_data()

    db_link = sqlite3.connect(r'sqlite_demo.db')
    db_csor = db_link.cursor()

    for film in films:
        # print(film[0])
        db_csor.execute(r"INSERT INTO FILMS_INFO VALUES('{:s}', '{:s}', '{:s}', '{:s}')".format(film[0], film[1], film[2], film[3]))

    db_csor.close()
    db_link.commit()
    db_link.close()


if __name__ == '__main__':
    insert_database()