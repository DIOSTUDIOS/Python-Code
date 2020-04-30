import openpyxl

sheets = []

def main():
    file = openpyxl.load_workbook("1图纸封皮(2).xlsx")
    # print(type(file))

    # sheets = file.get_sheet_names()
    print(len(file.sheetnames))

    for item in file.sheetnames:
        sheets.append(item)

    print(sheets[0])
    print(sheets[1])


if __name__ == '__main__':
    main()