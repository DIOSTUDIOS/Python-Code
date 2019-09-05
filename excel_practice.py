from openpyxl import Workbook


wb = Workbook()
# print(type(wb))
# print(wb.active)

# ws = wb.active
# ws.title = "example sheet"
# # print(wb.active)
# # print(ws.title)
# ws.sheet_properties.tabcolor = "1072BA"
# ws2 = wb.create_sheet("sheet 2")
# ws3 = wb.create_sheet("sheet 3")
# print(wb.sheetnames)

# for sheet in wb:
# 	print(sheet.title)

# ws4 = wb.copy_worksheet(wb["example sheet"])
# print(wb.sheetnames)

# wb.active = wb["sheet 2"]

# wb["Sheet"]["A4"] = 99
# print(wb["Sheet"]["A4"].row)
# print(wb["Sheet"]["A4"].column)
# print(wb["Sheet"]["A4"].value)

for number in range(1, 11):
	wb["Sheet"].cell(column="A", row=number, value=number)

# for number in range(1, 11):
# 	wb["sheet 2"].cell(column=2, row=number, value=number)

# for number in range(1, 11):
# 	wb["sheet 2"].cell(column=3, row=number, value=number)

# print(tuple(wb["sheet 2"].values))

# print(str(wb["sheet 2"]["C5"].value))

# for rows in wb["sheet 2"].values:
# 	print(rows)
	# for value in rows:
	# 	print(value)

# for cols in wb["sheet 2"].values:
# 	for value in cols:
# 		print(value)
# wb.remove(wb["sheet 3"])

# for col in range(1, 11):
# 	for row in range(1, 11):
# 		wb["sheet 2"].cell(column=col, row=row, value=(col + row))

wb.save(filename=r"example.xlsx")